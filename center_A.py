from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By

import pandas as pd
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SAVE_PATH = r"G:\내 드라이브\배민 자동화\중구A_실적.xlsx"
HTML_PATH = r"G:\내 드라이브\배민 자동화\중구A_dashboard.html"
REFRESH_MINUTES = 1

SET_RULES = {
    "월": [21, 20, 30, 29],
    "화": [21, 20, 30, 29],
    "수": [21, 20, 30, 29],
    "목": [21, 20, 30, 29],
    "금": [24, 21, 32, 33],
    "토": [31, 22, 36, 31],
    "일": [33, 22, 35, 30],
}

SONIC_TEAM = [
    "강우기", "강준호", "곽동욱", "권예준", "김도훈", "김동현", "김민", "김민승",
    "김보규", "김상일", "김용상", "김윤재", "김일수", "김지원", "김희경",
    "나원국", "노준우", "박경진", "박동혁", "박시우", "박재현", "박정권",
    "박진석", "박태우", "서정원", "손동국", "신정훈", "이국부", "이근화",
    "이상원", "이상호", "이승욱", "이재민", "임현철", "정창규", "진우일",
    "최순", "최준수", "최준호", "황정훈", "황주호", "박경보", "김장현",
    "길강호", "천용진", "이정빈", "김종찬", "성기창","조상홍","김정호","김장현"
]


def today_korean_weekday():
    now = datetime.now()
    if now.hour < 3 or (now.hour == 3 and now.minute < 10):
        now = now - timedelta(days=1)
    return ["월", "화", "수", "목", "금", "토", "일"][now.weekday()]


def click_page(driver, page_number):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)

    script = """
    const target = arguments[0];
    const buttons = Array.from(document.querySelectorAll('button'));
    const btn = buttons.find(b => b.innerText.trim() === target);
    if (btn) {
        btn.click();
        return true;
    }
    return false;
    """

    clicked = driver.execute_script(script, str(page_number))
    time.sleep(4)
    return clicked


def make_summary(df):
    weekday = today_korean_weekday()
    morning_rule, afternoon_rule, dinner_rule, night_rule = SET_RULES[weekday]

    total_complete = df["총완료"].sum()
    total_reject = df["거절"].sum()
    total_cancel = df["취소"].sum()
    total_delivery_cancel = df["배달취소"].sum()
    total_requests = total_complete + total_reject + total_cancel + total_delivery_cancel

    accept_rate = round((total_complete / total_requests * 100), 1) if total_requests else 0
    running_count = len(df[df["운행상태"].astype(str).str.contains("운행중", na=False)])

    morning = df["오전피크"].sum()
    afternoon = df["오후논피크"].sum()
    dinner = df["저녁피크"].sum()
    night = df["심야논피크"].sum()

    total_done = morning + afternoon + dinner + night
    total_target = (morning_rule + afternoon_rule + dinner_rule + night_rule) * 5

    return {
        "요일": weekday,
        "접속중": running_count,
        "총완료": total_complete,
        "전체요청": total_requests,
        "수락률": accept_rate,
        "오전피크": morning,
        "오전세트": round(morning / (morning_rule * 5), 2),
        "오후논피크": afternoon,
        "오후논피세트": round(afternoon / (afternoon_rule * 5), 2),
        "저녁피크": dinner,
        "저녁세트": round(dinner / (dinner_rule * 5), 2),
        "심야논피크": night,
        "심야세트": round(night / (night_rule * 5), 2),
        "세트달성률": round((total_done / total_target * 100), 1) if total_target else 0,
    }


def style_excel(file_path, summary):
    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "중구A현황"

    ws.insert_rows(1, 8)

    dark = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    blue = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    purple = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    header_fill = PatternFill(start_color="EAF3F8", end_color="EAF3F8", fill_type="solid")
    running_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    stopped_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:K1")
    ws["A1"] = f"중구A 실적 요약  |  {summary['요일']}요일 기준"
    ws["A1"].fill = dark
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    cards = [
        ("A3:B4", "접속중", summary["접속중"], green),
        ("C3:D4", "총완료", summary["총완료"], blue),
        ("E3:F4", "수락률", f"{summary['수락률']}%", yellow),
        ("G3:H4", "전체요청", summary["전체요청"], purple),
        ("I3:K4", "세트달성률", f"{summary['세트달성률']}%", red),
        ("A6:B7", "오전피크", f"{summary['오전피크']} / {summary['오전세트']}세트", blue),
        ("C6:D7", "오후논피크", f"{summary['오후논피크']} / {summary['오후논피세트']}세트", green),
        ("E6:F7", "저녁피크", f"{summary['저녁피크']} / {summary['저녁세트']}세트", yellow),
        ("G6:H7", "심야논피크", f"{summary['심야논피크']} / {summary['심야세트']}세트", purple),
    ]

    for cell_range, label, value, fill in cards:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(":")[0]]
        cell.value = f"{label}\n{value}"
        cell.fill = fill
        cell.font = Font(bold=True, size=15, color="111827")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in range(3, 8):
        ws.row_dimensions[row].height = 38

    table_header_row = 9
    ws.freeze_panes = f"A{table_header_row + 1}"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=table_header_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(color="222222")

    for cell in ws[table_header_row]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="1F4E5F")

    for row_num in range(table_header_row + 1, ws.max_row + 1):
        status = str(ws[f"B{row_num}"].value).strip()

        if "운행중" in status:
            row_fill = running_fill
            font = Font(color="006100", bold=True)
            ws.row_dimensions[row_num].hidden = False
        elif "운행 종료" in status:
            row_fill = stopped_fill
            font = Font(color="9C0006", bold=True)
            ws.row_dimensions[row_num].hidden = True
        else:
            row_fill = white
            font = Font(color="222222")
            ws.row_dimensions[row_num].hidden = False

        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = row_fill
            if col_num == 2:
                cell.font = font

    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True, color="7F6000")

    ws.row_dimensions[ws.max_row].hidden = False

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max(max_length + 3, 11)

    wb.save(file_path)


def make_dashboard_html(df):
    weekday = today_korean_weekday()
    morning_set, afternoon_set, dinner_set, night_set = SET_RULES[weekday]

    sonic_df = df[df["이름"].isin(SONIC_TEAM)].copy()
    number_df = df[~df["이름"].isin(SONIC_TEAM)].copy()

    total_complete = int(df["총완료"].sum())
    total_reject = int(df["거절"].sum())
    total_dispatch_cancel = int(df["취소"].sum())
    total_delivery_cancel = int(df["배달취소"].sum())

    total_requests = total_complete + total_reject + total_dispatch_cancel + total_delivery_cancel
    daily_accept_rate = round((total_complete / total_requests) * 100, 1) if total_requests else 0
    weekly_accept_rate = daily_accept_rate

    available_rejects = int(max((total_complete / 0.8) - total_requests, 0)) if total_complete else 0

    total_running = len(df[df["운행상태"].astype(str).str.contains("운행중", na=False)])
    sonic_running = len(sonic_df[sonic_df["운행상태"].astype(str).str.contains("운행중", na=False)])
    number_running = total_running - sonic_running

    updated = datetime.now().strftime("%Y.%m.%d %H:%M 기준")

    def make_peak_card(title, col, set_target):
        sonic_value = int(sonic_df[col].sum())
        number_value = int(number_df[col].sum())

        sonic_target = set_target * 4
        number_target = set_target

        total_value = sonic_value + number_value
        total_target = set_target * 5

        sonic_width = min(round((sonic_value / sonic_target) * 100, 1), 100) if sonic_target else 0
        number_width = min(round((number_value / number_target) * 100, 1), 100) if number_target else 0

        return f"""
        <div class="peak-card">
            <div class="peak-title">
                <div class="mini-logo">S</div>
                <span>{title} ({total_value}/{total_target})</span>
            </div>

            <div class="bar-row">
                <div class="bar-label">소닉</div>
                <div class="bar-wrap">
                    <div class="bar-fill" style="width:{sonic_width}%"></div>
                    <div class="bar-text">{sonic_value}/{sonic_target}</div>
                </div>
            </div>

            <div class="bar-row">
                <div class="bar-label">넘버</div>
                <div class="bar-wrap">
                    <div class="bar-fill" style="width:{number_width}%"></div>
                    <div class="bar-text">{number_value}/{number_target}</div>
                </div>
            </div>
        </div>
        """

    peak_cards = "".join([
        make_peak_card("오전피크", "오전피크", morning_set),
        make_peak_card("오후논피크", "오후논피크", afternoon_set),
        make_peak_card("저녁피크", "저녁피크", dinner_set),
        make_peak_card("심야논피크", "심야논피크", night_set),
    ])

    html = f"""
<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="refresh" content="60">
<title>슈퍼소닉 중구A</title>
<style>
* {{ box-sizing: border-box; }}
body {{
    margin: 0;
    padding: 24px 14px 40px;
    background: #ffffff;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Noto Sans KR", sans-serif;
    color: #111;
}}
.wrap {{ max-width: 980px; margin: 0 auto; }}
.logo {{
    width: 70px; height: 70px; margin: 0 auto 12px; border-radius: 50%;
    background: #ff1630; color: white; display: flex; align-items: center; justify-content: center;
    font-size: 42px; font-weight: 900;
}}
.title {{ text-align: center; font-size: 42px; font-weight: 900; margin: 0; }}
.updated {{ margin-top: 10px; text-align: center; color: #666; font-size: 16px; font-weight: 700; margin-bottom: 26px; }}
.summary {{
    display: grid; grid-template-columns: repeat(2, 1fr);
    border: 3px solid #ff1630; border-radius: 18px; overflow: hidden; margin-bottom: 28px;
}}
.summary-item {{
    min-height: 108px; display: flex; flex-direction: column; align-items: center; justify-content: center;
    border-right: 2px solid #ff1630; border-bottom: 2px solid #ff1630;
}}
.summary-item:nth-child(2n) {{ border-right: none; }}
.summary-item:nth-last-child(-n+2) {{ border-bottom: none; }}
.summary-label {{ font-size: 20px; font-weight: 900; margin-bottom: 8px; }}
.summary-value {{ font-size: 32px; font-weight: 900; }}
.red {{ color: #e60012; }}
.blue {{ color: #2563eb; }}
.green {{ color: #079b24; }}
.yellow {{ color: #d97706; }}
.peaks {{ display: grid; grid-template-columns: 1fr; gap: 18px; }}
.peak-card {{ border: 3px solid #ff1630; border-radius: 22px; padding: 22px; }}
.peak-title {{ display: flex; align-items: center; gap: 12px; font-size: 24px; font-weight: 900; margin-bottom: 22px; }}
.mini-logo {{
    width: 38px; height: 38px; border-radius: 50%; background: #ff1630; color: white;
    display: flex; align-items: center; justify-content: center; font-size: 22px; font-weight: 900;
}}
.bar-row {{ display: grid; grid-template-columns: 70px 1fr; gap: 12px; align-items: center; margin-bottom: 16px; }}
.bar-label {{ font-size: 22px; font-weight: 900; }}
.bar-wrap {{ position: relative; height: 28px; background: #ffd1d8; border-radius: 999px; overflow: hidden; }}
.bar-fill {{ height: 100%; background: linear-gradient(90deg, #ff3b50, #ff1630); }}
.bar-text {{
    position: absolute; inset: 0; display: flex; align-items: center; justify-content: center;
    font-size: 18px; font-weight: 900;
}}
.footer {{ margin-top: 24px; text-align: center; color: #666; font-size: 18px; font-weight: 700; }}
</style>
</head>

<body>
<div class="wrap">
    <div class="logo">S</div>
    <h1 class="title">슈퍼소닉 중구A</h1>
    <div class="updated">🕘 {updated}</div>

    <div class="summary">
        <div class="summary-item"><div class="summary-label">총완료</div><div class="summary-value red">{total_complete:,}</div></div>
        <div class="summary-item"><div class="summary-label">총거절</div><div class="summary-value red">{total_reject:,}</div></div>
        <div class="summary-item"><div class="summary-label">배차취소</div><div class="summary-value red">{total_dispatch_cancel:,}</div></div>
        <div class="summary-item"><div class="summary-label">배달취소</div><div class="summary-value red">{total_delivery_cancel:,}</div></div>
        <div class="summary-item"><div class="summary-label">주간수락률</div><div class="summary-value yellow">{weekly_accept_rate}%</div></div>
        <div class="summary-item"><div class="summary-label">당일수락률</div><div class="summary-value blue">{daily_accept_rate}%</div></div>
        <div class="summary-item"><div class="summary-label">거절가능</div><div class="summary-value yellow">{available_rejects}개</div></div>
        <div class="summary-item"><div class="summary-label">전체접속</div><div class="summary-value green">{total_running}명</div></div>
        <div class="summary-item"><div class="summary-label">소닉팀</div><div class="summary-value green">{sonic_running}명</div></div>
        <div class="summary-item"><div class="summary-label">넘버팀</div><div class="summary-value green">{number_running}명</div></div>
    </div>

    <div class="peaks">
        {peak_cards}
    </div>

    <div class="footer">↻ 1분마다 자동 갱신 중...</div>
</div>
</body>
</html>
"""

    Path(HTML_PATH).write_text(html, encoding="utf-8")
    print("중구A 대시보드 생성 완료")
    print(HTML_PATH)


def save_excel(results):
    if not results:
        print("찾은 기사가 없습니다.")
        return

    df = pd.DataFrame(results).drop_duplicates()

    df["정렬순서"] = df["운행상태"].apply(lambda x: 0 if "운행중" in str(x) else 1)
    df = df.sort_values(
        by=["정렬순서", "총완료", "이름"],
        ascending=[True, False, True]
    )
    df = df.drop(columns=["정렬순서"])

    summary = make_summary(df)

    total_row = {
        "이름": "합계",
        "운행상태": "",
        "총완료": df["총완료"].sum(),
        "거절": df["거절"].sum(),
        "취소": df["취소"].sum(),
        "배달취소": df["배달취소"].sum(),
        "오전피크": df["오전피크"].sum(),
        "오후논피크": df["오후논피크"].sum(),
        "저녁피크": df["저녁피크"].sum(),
        "심야논피크": df["심야논피크"].sum(),
        "아이디": "",
    }

    excel_df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    excel_df.to_excel(SAVE_PATH, index=False)
    style_excel(SAVE_PATH, summary)
    make_dashboard_html(df)

    print("\n중구A 엑셀 저장 완료")
    print(SAVE_PATH)
    print(f"요일 기준: {summary['요일']}요일")
    print("※ 00:00~03:09까지는 전날 요일 기준입니다.")


def collect_current_pages(driver):
    results = []

    for page in [1, 2, 3, 4]:
        print(f"\n--- {page}페이지 확인 중 ---")

        clicked = click_page(driver, page)

        if not clicked:
            print(f"{page}페이지 버튼을 못 찾았습니다.")

        rows = driver.find_elements(By.TAG_NAME, "tr")

        for row in rows:
            text = row.text.strip()

            if not text:
                continue

            data = text.split("\n")

            try:
                if len(data) >= 11 and data[0] != "합계":
                    results.append({
                        "이름": data[0],
                        "운행상태": data[1],
                        "총완료": int(data[3]),
                        "거절": int(data[4]),
                        "취소": int(data[5]),
                        "배달취소": int(data[6]),
                        "오전피크": int(data[7]),
                        "오후논피크": int(data[8]),
                        "저녁피크": int(data[9]),
                        "심야논피크": int(data[10]),
                        "아이디": data[11],
                    })

                    team = "소닉팀" if data[0] in SONIC_TEAM else "넘버팀"
                    print(f"저장: {data[0]} / {team}")

            except Exception as e:
                pass

    return results


driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://deliverycenter.baemin.com")

print("중구A 권역으로 로그인 후 기사목록 화면까지 이동하세요.")
input("기사목록 화면이면 엔터 누르세요...")

while True:
    print("\n화면 새로고침 중...")
    driver.refresh()
    time.sleep(6)

    results = collect_current_pages(driver)
    save_excel(results)

    print(f"\n{REFRESH_MINUTES}분 뒤 다시 갱신합니다.")
    print("종료하려면 CMD 창에서 Ctrl + C 누르세요.")

    time.sleep(REFRESH_MINUTES * 60)
